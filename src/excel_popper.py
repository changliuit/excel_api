import pathlib
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl import Workbook
import requests

from all_exceptions import InputError
from data_classes import ApiRequest, Report


class ExcelPopper:
    """Read an excel, query api, and write response to another excel"""
    def __init__(self, config_path: Optional[str] = None, output_filepath: Optional[str] = None):
        self.api_url: str = "https://covid-api.com/api/reports"
        self.config_path: str = config_path or 'input.config'
        self.output_filepath: str = output_filepath or 'output.xlsx'

    def read_excel(self) -> List[ApiRequest]:
        """return all the combinations of country and date from the excel file in config"""
        print(f"Opening {self.config_path}")
        workbook = load_workbook(filename=self.config_path)
        result = []
        for row in workbook.active.rows:
            if row[0].value is not None:
                request = ApiRequest(
                    iso=row[0].value,
                    date=row[1].value,
                )
                result.append(request)
        print(f"{len(result)} lines loaded from file")
        return result

    def query_api_for_report(self, api_request: ApiRequest):
        response = requests.get(self.api_url, params=api_request.__dict__)
        return response

    def get_report_to_write(self, api_requests: List[ApiRequest]) -> List[Report]:
        """query api and return reports in needed format"""
        results = []
        for i, req in enumerate(api_requests):
            print(f"Querying api for line {i + 1}...")
            response = self.query_api_for_report(req)
            if response.status_code == 200:
                if response.json().get('data'):
                    reports = [
                        Report(
                            date=item['date'],
                            iso=item['region']['iso'],
                            num_confirmed=item['confirmed'],
                            num_deaths=item['deaths'],
                            num_recovered=item['recovered']
                        ) for item in response.json()['data']
                    ]
                    results.extend(reports)
            else:
                raise InputError(f'Input data not valid. Error response: {response.json()}')
        return results

    def write_excel(self, reports: List[Report]):
        """Write a list of reports to Excel"""
        print(f"Saving reports to {self.output_filepath}...")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Covid Numbers"
        for report in reports:
            ws1.append(list(report.__dict__.values()))
        wb.save(filename=self.output_filepath)
        print(f"Saved to {self.output_filepath}")
